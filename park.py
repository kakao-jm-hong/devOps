import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import Border, Side
from openpyxl.cell.cell import MergedCell

# 입력 값
user_input = input("날짜를 입력하세요 (예: 2025-08-12, 생략 시 오늘 날짜): ").strip()
try:
    date = datetime.strptime(user_input, "%Y-%m-%d").strftime("%Y-%m-%d") if user_input else date.today().strftime("%Y-%m-%d")
except ValueError:
    print("잘못된 형식입니다. 오늘 날짜로 진행합니다.")
    date = date.today().strftime("%Y-%m-%d")
# 시작 시간 입력
start_hour = int(input("근무 시작 시간을 입력하세요 (8 또는 9): "))
if start_hour == 8:
    end_hour = 17
elif start_hour == 9:
    end_hour = 18
else:
    raise ValueError("시작 시간은 8 또는 9만 입력 가능합니다.")
비고내용 = "육전 연결 중,"

# 각 열 데이터 구성
rows = []
for hour in range(start_hour, end_hour + 1):
    row = {
        "요목 구분1": "기사",
        "요목 구분2": "기사일반",
        "일자": date,
        "시": hour,
        "분": 0,
        "합정연혁 포함여부": "N",
        "기사제목": "",
        "기사내용": 비고내용
    }
    rows.append(row)

# DataFrame 생성
df = pd.DataFrame(rows)

# 컬럼 순서 조정
columns = ["요목 구분1", "요목 구분2", "일자", "시", "분", "합정연혁 포함여부", "기사제목", "기사내용"]
df = df[columns]

# 작성자 정보 추가 (엑셀 오른쪽 하단 셀)
author_row = pd.DataFrame([{col: "육전 연결 중," if col == "기사내용" else "" for col in columns}])
df = pd.concat([df, author_row], ignore_index=True)

# Excel로 저장
output_path = f"{date}_기사일지.xlsx"
# Disable automatic header insertion and start data from row 4
df.to_excel(output_path, index=False, header=False, startrow=3)

# 엑셀 파일 불러와서 스타일 적용
wb = load_workbook(output_path)
ws = wb.active
fill_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 노란색

# 병합 셀 처리
ws.merge_cells("A1:B2")  # 요목 전체 병합
ws.merge_cells("C1:C2")
ws.merge_cells("C1:E2")
ws.merge_cells("F1:F3")  # 합정연혁 포함여부
ws.merge_cells("G1:G3")  # 기사제목
ws.merge_cells("H1:H3")  # 기사내용

ws.merge_cells("A3")  # 요목 구분1
ws.merge_cells("B3")  # 요목 구분2
ws.merge_cells("C3")  # 일자
ws.merge_cells("D3")  # 시
ws.merge_cells("E3")  # 분

# Insert text into the subheaders under "요목"
ws["A3"] = "요목 구분1"
ws["B3"] = "요목 구분2"
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
ws["A3"].fill = fill_color
ws["B3"].fill = fill_color

ws.merge_cells("C1:E1")  # 기사일시
ws.merge_cells("C2:E2")  # 시작일자

# 병합된 셀에 값 입력 시, 반드시 병합된 영역의 첫 번째 셀(좌상단)에만 값을 넣어야 함
ws["A1"] = "요목"
ws["C1"] = "기사일시"
ws["C1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["C3"] = "일자"
ws["D3"] = "시"
ws["E3"] = "분"

# Add red triangle indicators (comments) to A1 and C1
from openpyxl.comments import Comment
ws["A1"].comment = Comment("필수 입력 항목입니다.", "System")
ws["C1"].comment = Comment("필수 입력 항목입니다.", "System")

ws["F1"] = "함정연혁 포함여부"
ws["G1"] = "기사제목"
ws["H1"] = "기사내용"
ws["H1"].comment = Comment("필수 입력 항목입니다.", "System")

# 스타일 지정
for col in ["A", "B", "F", "G", "H"]:
    for row in range(1, 4):
        ws[f"{col}{row}"].fill = fill_color
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")


# Define thin border style
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)



# Apply border to all header cells in rows 1 to 3 (A1:H3)
for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=8):
    for cell in row:
        cell.border = thin_border

# Explicitly re-apply border to merged cells in row 1 to ensure border is visible
ws["A1"].border = thin_border
ws["B1"].border = thin_border
ws["C1"].border = thin_border
ws["D1"].border = thin_border
ws["E1"].border = thin_border
ws["F1"].border = thin_border
ws["G1"].border = thin_border
ws["H1"].border = thin_border


# 컬럼 너비 직접 지정
ws.column_dimensions["A"].width = 10  # 요목 구분 1
ws.column_dimensions["B"].width = 10  # 요목 구분 2
ws.column_dimensions["C"].width = 14  # 일자
ws.column_dimensions["D"].width = 5   # 시
ws.column_dimensions["E"].width = 5   # 분
ws.column_dimensions["F"].width = 20  # 함정연혁 포함여부
ws.column_dimensions["G"].width = 20  # 기사제목
ws.column_dimensions["H"].width = 100  # 기사내용


# I30 셀에 "박명수" 입력
ws["I30"] = "박명수"

wb.save(output_path)

print(f"{output_path} 파일 생성 완료!")
